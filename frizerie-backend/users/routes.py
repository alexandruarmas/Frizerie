from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from typing import Dict, Any, List, Optional
from datetime import timedelta
import csv
from fastapi.responses import StreamingResponse
from io import StringIO
import openpyxl
from io import BytesIO

from config.database import get_db
from auth.dependencies import get_current_user, get_current_admin
from config.settings import get_settings
from auth.services import create_access_token, create_refresh_token
from . import services
from . import models
from validation import (
    UserCreate,
    UserUpdate,
    UserResponse,
    validate_password_strength
)
from validation.schemas import (
    LoyaltyStatus,
    LoyaltyRewardResponse,
    LoyaltyRedemptionResponse,
    LoyaltyRedemptionUpdate,
    LoyaltyPointsHistoryResponse,
    ReferralProgramResponse,
    ReferralProgramCreate,
    LoyaltyRewardCreate,
    LoyaltyRewardUpdate,
)
from .models import AuditLog

router = APIRouter(prefix="/users", tags=["Users"])
settings = get_settings()

@router.post("/", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    db: Session = Depends(get_db)
):
    """Create a new user."""
    try:
        # Validate password strength
        validate_password_strength(user_data.password)
        
        # Create user
        user = services.create_user(db, user_data)
        
        # Create tokens
        access_token = create_access_token(
            data={"sub": user.email},
            expires_delta=timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE_MINUTES)
        )
        refresh_token = create_refresh_token(
            data={"sub": user.email},
            expires_delta=timedelta(days=settings.REFRESH_TOKEN_EXPIRE_DAYS)
        )
        
        return {
            **user.to_dict(),
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer"
        }
    except Exception as e:
        print("ERROR IN CREATE_USER ROUTE:", str(e))
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

@router.get("/me", response_model=UserResponse)
async def get_current_user_info(
    current_user: models.User = Depends(get_current_user)
):
    """Get current user information."""
    return current_user

@router.put("/me", response_model=UserResponse)
async def update_current_user(
    user_data: UserUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update current user information."""
    return services.update_user(db, current_user.id, user_data.dict(exclude_unset=True))

@router.get("/vip-tiers", response_model=List[Dict[str, Any]])
async def get_vip_tiers(
    db: Session = Depends(get_db)
):
    """Get all VIP tiers."""
    tiers = services.get_all_vip_tiers(db)
    return [tier.to_dict() for tier in tiers]

@router.get("/vip-tiers/{tier_name}", response_model=Dict[str, Any])
async def get_vip_tier(
    tier_name: str,
    db: Session = Depends(get_db)
):
    """Get information about a specific VIP tier."""
    tier = services.get_vip_tier_info(db, tier_name)
    if not tier:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="VIP tier not found"
        )
    return tier.to_dict()

@router.post("/vip-tiers", response_model=Dict[str, Any])
async def create_vip_tier(
    tier_data: Dict[str, Any],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin)
):
    """Create a new VIP tier (admin only)."""
    return services.create_vip_tier(
        db,
        tier_data["name"],
        tier_data["min_points"],
        tier_data["perks"]
    ).to_dict()

@router.put("/vip-tiers/{tier_name}", response_model=Dict[str, Any])
async def update_vip_tier(
    tier_name: str,
    tier_data: Dict[str, Any],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin)
):
    """Update a VIP tier (admin only)."""
    return services.update_vip_tier(db, tier_name, tier_data).to_dict()

@router.post("/me/loyalty-points", response_model=UserResponse)
async def add_loyalty_points(
    points: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Add loyalty points to current user."""
    if points <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Points must be positive"
        )
    return services.add_loyalty_points(db, current_user.id, points)

@router.get("/me/loyalty-status", response_model=LoyaltyStatus)
async def get_loyalty_status(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get detailed loyalty status for current user."""
    return services.get_loyalty_status(db, current_user.id)

@router.get("/me/loyalty/rewards", response_model=List[LoyaltyRewardResponse])
async def get_available_rewards(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get available rewards for current user."""
    return services.get_available_rewards(db, current_user.id)

@router.post("/me/loyalty/rewards/{reward_id}/redeem", response_model=LoyaltyRedemptionResponse)
async def redeem_reward(
    reward_id: int,
    booking_id: Optional[int] = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Redeem a loyalty reward."""
    return services.redeem_reward(db, current_user.id, reward_id, booking_id)

@router.get("/me/loyalty/points-history", response_model=List[LoyaltyPointsHistoryResponse])
async def get_points_history(
    skip: int = 0,
    limit: int = 10,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get user's points history."""
    return services.get_points_history(db, current_user.id, skip, limit)

@router.post("/me/loyalty/refer", response_model=ReferralProgramResponse)
async def create_referral(
    referral_data: ReferralProgramCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new referral."""
    return services.create_referral(db, current_user.id, referral_data.referred_email)

@router.post("/me/loyalty/referrals/{referral_id}/complete", response_model=ReferralProgramResponse)
async def complete_referral(
    referral_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Complete a referral and award points."""
    # Verify the referral belongs to the current user
    referral = db.query(models.ReferralProgram).filter(
        models.ReferralProgram.id == referral_id,
        models.ReferralProgram.referrer_id == current_user.id
    ).first()
    
    if not referral:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Referral not found or not authorized"
        )
    
    return services.complete_referral(db, referral_id)

# Admin routes for managing loyalty program
@router.post("/admin/loyalty/rewards", response_model=LoyaltyRewardResponse)
async def create_reward(
    reward_data: LoyaltyRewardCreate,
    current_user: models.User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Create a new loyalty reward (admin only)."""
    reward = models.LoyaltyReward(**reward_data.dict())
    db.add(reward)
    db.commit()
    db.refresh(reward)
    return reward

@router.put("/admin/loyalty/rewards/{reward_id}", response_model=LoyaltyRewardResponse)
async def update_reward(
    reward_id: int,
    reward_data: LoyaltyRewardUpdate,
    current_user: models.User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update a loyalty reward (admin only)."""
    reward = db.query(models.LoyaltyReward).filter(
        models.LoyaltyReward.id == reward_id
    ).first()
    
    if not reward:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Reward not found"
        )
    
    for field, value in reward_data.dict(exclude_unset=True).items():
        setattr(reward, field, value)
    
    db.commit()
    db.refresh(reward)
    return reward

@router.get("/admin/loyalty/redemptions", response_model=List[LoyaltyRedemptionResponse])
async def get_all_redemptions(
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None,
    current_user: models.User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Get all loyalty redemptions (admin only)."""
    query = db.query(models.LoyaltyRedemption)
    
    if status:
        query = query.filter(models.LoyaltyRedemption.status == status)
    
    redemptions = query.order_by(
        models.LoyaltyRedemption.redeemed_at.desc()
    ).offset(skip).limit(limit).all()
    
    return redemptions

@router.put("/admin/loyalty/redemptions/{redemption_id}", response_model=LoyaltyRedemptionResponse)
async def update_redemption_status(
    redemption_id: int,
    redemption_data: LoyaltyRedemptionUpdate,
    current_user: models.User = Depends(get_current_admin),
    db: Session = Depends(get_db)
):
    """Update redemption status (admin only)."""
    redemption = db.query(models.LoyaltyRedemption).filter(
        models.LoyaltyRedemption.id == redemption_id
    ).first()
    
    if not redemption:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Redemption not found"
        )
    
    for field, value in redemption_data.dict(exclude_unset=True).items():
        setattr(redemption, field, value)
    
    db.commit()
    db.refresh(redemption)
    return redemption

@router.get("/all", response_model=List[UserResponse])
async def get_all_users(
    db: Session = Depends(get_db)
):
    """Get all users."""
    try:
        users = services.get_all_users(db)
        return users
    except Exception as e:
        print("ERROR IN GET_ALL_USERS:", str(e))
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )

# Helper: Block access if terms_accepted is False
async def check_terms_accepted(current_user: models.User = Depends(get_current_user)):
    if not current_user.terms_accepted:
        raise HTTPException(status_code=403, detail="Trebuie să accepți Termenii și Politica de Confidențialitate pentru a folosi serviciul.")
    return current_user

@router.put("/me/withdraw-consent", response_model=Dict[str, str])
async def withdraw_consent(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    current_user.terms_accepted = False
    current_user.terms_accepted_at = None
    db.commit()
    return {"detail": "Consimțământul a fost retras. Nu vei mai putea folosi serviciul până nu accepți din nou termenii."}

@router.get("/me/export", response_model=Dict[str, Any])
async def export_user_data(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_terms_accepted)
):
    bookings = db.query(models.Booking).filter_by(user_id=current_user.id).all() if hasattr(models, 'Booking') else []
    payments = db.query(models.Payment).filter_by(user_id=current_user.id).all() if hasattr(models, 'Payment') else []
    notifications = db.query(models.Notification).filter_by(user_id=current_user.id).all() if hasattr(models, 'Notification') else []
    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action="export",
        resource_type="user",
        resource_id=current_user.id,
        details={"info": "User data export"}
    )
    db.add(audit)
    db.commit()
    return {
        "profile": current_user.to_dict(),
        "bookings": [b.to_dict() for b in bookings],
        "payments": [p.to_dict() for p in payments],
        "notifications": [n.to_dict() for n in notifications],
    }

@router.get("/me/export-csv")
async def export_user_data_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_terms_accepted)
):
    bookings = db.query(models.Booking).filter_by(user_id=current_user.id).all() if hasattr(models, 'Booking') else []
    payments = db.query(models.Payment).filter_by(user_id=current_user.id).all() if hasattr(models, 'Payment') else []
    notifications = db.query(models.Notification).filter_by(user_id=current_user.id).all() if hasattr(models, 'Notification') else []
    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action="export_csv",
        resource_type="user",
        resource_id=current_user.id,
        details={"info": "User data export CSV"}
    )
    db.add(audit)
    db.commit()
    # Prepare CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Field", "Value"])
    for k, v in current_user.to_dict().items():
        writer.writerow([k, v])
    output.write("\nBookings\n")
    for b in bookings:
        writer.writerow(["Booking"] + list(b.to_dict().values()))
    output.write("\nPayments\n")
    for p in payments:
        writer.writerow(["Payment"] + list(p.to_dict().values()))
    output.write("\nNotifications\n")
    for n in notifications:
        writer.writerow(["Notification"] + list(n.to_dict().values()))
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=user_data.csv"})

@router.get("/me/export-xlsx")
async def export_user_data_xlsx(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_terms_accepted)
):
    bookings = db.query(models.Booking).filter_by(user_id=current_user.id).all() if hasattr(models, 'Booking') else []
    payments = db.query(models.Payment).filter_by(user_id=current_user.id).all() if hasattr(models, 'Payment') else []
    notifications = db.query(models.Notification).filter_by(user_id=current_user.id).all() if hasattr(models, 'Notification') else []

    wb = openpyxl.Workbook()
    ws_profile = wb.active
    ws_profile.title = "Profile"
    for k, v in current_user.to_dict().items():
        ws_profile.append([k, v])

    ws_bookings = wb.create_sheet("Bookings")
    for b in bookings:
        ws_bookings.append(list(b.to_dict().values()))

    ws_payments = wb.create_sheet("Payments")
    for p in payments:
        ws_payments.append(list(p.to_dict().values()))

    ws_notifications = wb.create_sheet("Notifications")
    for n in notifications:
        ws_notifications.append(list(n.to_dict().values()))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=user_data.xlsx"})

@router.get("/me/audit-logs", response_model=List[dict])
async def get_my_audit_logs(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_terms_accepted)
):
    logs = db.query(AuditLog).filter(AuditLog.user_id == current_user.id).order_by(AuditLog.created_at.desc()).all()
    return [log.__dict__ for log in logs]

@router.get("/me/third-parties", response_model=List[dict])
async def get_third_parties_info(
    current_user: models.User = Depends(check_terms_accepted)
):
    third_parties = [
        {"name": "Stripe", "purpose": "Procesare plăți", "link": "https://stripe.com/privacy"},
        {"name": "Twilio", "purpose": "Trimitere SMS", "link": "https://www.twilio.com/legal/privacy"},
        {"name": "Sentry", "purpose": "Monitorizare erori", "link": "https://sentry.io/privacy/"},
        # Adaugă și alți furnizori dacă ai
    ]
    return third_parties

@router.delete("/me", response_model=Dict[str, str])
async def delete_user(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(check_terms_accepted)
):
    # Audit log
    audit = AuditLog(
        user_id=current_user.id,
        action="delete",
        resource_type="user",
        resource_id=current_user.id,
        details={"info": "User account deleted"}
    )
    db.add(audit)
    # Delete related data
    if hasattr(models, 'Booking'):
        db.query(models.Booking).filter_by(user_id=current_user.id).delete()
    if hasattr(models, 'Payment'):
        db.query(models.Payment).filter_by(user_id=current_user.id).delete()
    if hasattr(models, 'Notification'):
        db.query(models.Notification).filter_by(user_id=current_user.id).delete()
    db.delete(current_user)
    db.commit()
    return {"detail": "Contul și toate datele asociate au fost șterse."} 